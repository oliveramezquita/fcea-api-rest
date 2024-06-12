from fcea_monitoreo.utils import get_collection
from api.helpers.http_responses import not_found, error
from urllib.parse import parse_qs
from django.http import HttpResponse
from bson import ObjectId
from datetime import datetime, timedelta
from api.constants import IGNORE_KEYS, ARRAYS_VALUES, EXCEL_HEADER
from fcea_monitoreo.settings import BASE_URL
import json
import openpyxl
import time


class GetSitesUseCase:
    def __init__(self, request):
        params = parse_qs(request.META['QUERY_STRING'])

        # project
        self.project = params['project'][0] if 'project' in params else None
        # monitoring period
        self.monitoring_period = str(params['monitoring_period'][0]).split(
        ) if 'monitoring_period' in params else []
        # year
        self.year = self.monitoring_period[1] if len(
            self.monitoring_period) > 0 else None
        # month
        self.month = self.monitoring_period[0] if len(
            self.monitoring_period) > 0 else None
        # season
        self.season = params['season'][0] if 'season' in params else None
        # state
        self.state = params['state'][0] if 'state' in params else None
        # institution
        self.institution = params['institution'][0] if 'institution' in params else None
        # dates
        self.dates = params['dates'][0] if 'dates' in params else None

    def execute(self):
        try:
            filters = {}
            if self.project:
                project = get_collection(
                    'projects', {
                        'name': self.project,
                        'year': int(self.year),
                        'month': self.month,
                        'season': self.season,
                        '_deleted': False,
                    })
                if project:
                    filters['project_id'] = project[0]['_id']
                    if self.state:
                        filters['estado'] = self.state
                    if self.institution:
                        filters['institucion'] = self.institution
                    if self.dates:
                        filters['fecha'] = self.filter_by_dates()
            sites = self.get_sites(filters) if len(filters) > 0 else []
            if len(sites) > 0:
                self.create_xls_file(sites)
                return HttpResponse(json.dumps(sites), content_type='application/json')
            return not_found("No existen sitios dados de alta hasta el momento")
        except Exception as e:
            return error(e.args[0])

    def filter_by_dates(self):
        dates = self.dates.split(" to ")
        lt = datetime.strptime(dates[-1], '%Y-%m-%d')
        return {'$gte': datetime.strptime(dates[0], '%Y-%m-%d'), '$lt': lt + timedelta(days=1)}

    def get_site_filters(self):
        try:
            projects, monitoring_periods = self.get_filter_projects()
            index = next((idx for idx, p in enumerate(
                projects) if p["name"] == self.project and p["year"] == int(self.year) and p["month"] == self.month and p["season"] == self.season), 0)
            sites = get_collection(
                'sites', {'cuenca': projects[index]['name']}) if len(projects) > 0 else []
            if sites:
                states = list(set(p['estado'] for p in sites))
                institution = list(set(p['institucion'] for p in sites))
                seasons = list(set(p['temporada'] for p in sites))
                resp = {
                    'default_project': {
                        'name': projects[index]['name'],
                        'monitoring_period': monitoring_periods[index],
                        'season': projects[index]['season'],
                        'geojson_data': self._get_geojson_file(projects[index]['name']),
                        'institutions': self.get_institutions(projects[index]['name'], projects[index]['institutions']) if 'institutions' in projects[index] else [],
                        'excel_file': f"{BASE_URL}/media/files/{projects[index]['name']}_{projects[index]['season']}{str(monitoring_periods[index]).replace(' ','')}.xlsx",
                    },
                    'projects': self.get_disctint_projects(projects),
                    'monitoring_periods': self.get_filter_by_project(projects[index]['name']),
                    'states': states,
                    'institution': institution,
                    'seasons': seasons,
                }
                return HttpResponse(json.dumps(resp), content_type='application/json')
            return not_found("No existen filtros actualmente")
        except Exception as e:
            return error(e.args[0])

    def get_filter_projects(self):
        project_filter = {'name': self.project,
                          '_deleted': False} if self.project else {'_deleted': False}
        projects = get_collection('projects', project_filter)
        project_list = []
        monitoring_periods_list = []
        for project in projects:
            sites = get_collection(
                'sites', {'project_id': ObjectId(project['_id'])})
            if sites:
                project_list.append(project)
                monitoring_periods_list.append(
                    f"{project['month']} {project['year']}")

        return project_list, monitoring_periods_list

    def _get_geojson_file(self, cuenca_name):
        cuenca = get_collection('basins', {'name': cuenca_name})
        if cuenca:
            return cuenca[0]['geojson_file']
        return None

    def get_disctint_projects(self, projects):
        distinct_list_projects = []
        for project in projects:
            if project['name'] not in distinct_list_projects:
                distinct_list_projects.append(project['name'])
        return distinct_list_projects

    def get_filter_by_project(self, project_name):
        projects = get_collection(
            'projects', {'name': project_name, '_deleted': False})
        monitoring_period_list = []
        for project in projects:
            monitoring_period_list.append(
                f"{project['month']} {project['year']}")
        return monitoring_period_list

    def get_sites(self, filters):
        sites = get_collection('sites', filters)
        resp_sites = []
        for site in sites:
            del site['create_date']
            site['_id'] = str(site['_id'])
            site['project_id'] = str(site['project_id'])
            site['sitio_referencia_id'] = str(
                site['sitio_referencia_id']) if site['sitio_referencia_id'] else None
            site['reference_site_scores'] = self.get_reference_site_scores(
                site['sitio_referencia_id'])
            site['user_id'] = str(site['user_id'])
            site['fecha'] = site['fecha'].isoformat()
            resp_sites.append(site)
        return resp_sites

    def get_reference_site_scores(self, sitio_referencia_id):
        sitio_referencia = get_collection(
            'sites', {
                '_id': ObjectId(sitio_referencia_id),
                'es_sitio_referencia': True
            })
        if sitio_referencia:
            return {
                'nombre_sitio': sitio_referencia[0]['nombre_sitio'],
                'ph': sitio_referencia[0]['ph'],
                'temperatura_agua': sitio_referencia[0]['temperatura_agua'],
                'temperatura_ambiental': sitio_referencia[0]['temperatura_ambiental'],
                'turbidez': sitio_referencia[0]['turbidez'],
                'nitratos': sitio_referencia[0]['nitratos'],
                'amonio': sitio_referencia[0]['amonio'],
                'ortofosfatos': sitio_referencia[0]['ortofosfatos'],
                'saturacion': sitio_referencia[0]['saturacion'],
                'oxigeno_disuelto': sitio_referencia[0]['oxigeno_disuelto'],
            }
        return None

    def get_institutions(self, basin_name, institutions):
        data = []
        basin = get_collection('basins', {'name': basin_name})
        if basin and 'institutions' in basin[0]:
            institution_list = institutions.split(',')
            for institution in institution_list:
                inst = next(
                    (i for i in basin[0]['institutions'] if i['name'] == institution), None)
                if inst:
                    data.append(inst)
        return data

    def create_xls_file(self, sites):
        # header = self._get_key_list(sites)
        header = EXCEL_HEADER
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(header)
        filename = ''
        for idx, s in enumerate(sites):
            column = 1
            row = idx + 2
            filename = f'{s["cuenca"]}_{s["temporada"]}{s["mes"]}{s["anio"]}'
            for key, value in s.items():
                if key in ARRAYS_VALUES:
                    v = ''
                    if key == 'macroinvertebrados':
                        v = ', '.join(
                            f'{m["familia"]}: {m["puntaje"]}' for m in value)
                    if key == 'secciones':
                        v = ', '.join(
                            f'ancho: {s["width"]} - profundidad: {s["depth"]}' for s in value)
                    if key == 'archivos_adjuntos':
                        v = ', '.join(aa for aa in value)
                    sheet.cell(row=row, column=column, value=v)
                    column = column + 1
                elif key not in IGNORE_KEYS:
                    if key == 'user_id':
                        sheet.cell(row=row, column=column,
                                   value=self._get_user_name(value))
                    elif key == 'sitio_referencia_id':
                        v = next(
                            (srf['nombre_sitio'] for srf in sites if srf['_id'] == value), 'N/A')
                        sheet.cell(row=row, column=column, value=v)
                    else:
                        sheet.cell(row=row, column=column, value=value)
                    column = column + 1
                if key == 'scores':
                    sheet.cell(row=row, column=column,
                               value=value['total'][3])
                    sheet.cell(row=row, column=column+1,
                               value=value['interpretation'][0])
                    sheet.cell(row=row, column=column+2,
                               value=value['interpretation'][1])

        wb.save(f'media/files/{filename}.xlsx')

    def _get_key_list(self, sites):
        list_keys = []
        for s in sites:
            for key, _ in s.items():
                if key not in IGNORE_KEYS and key not in list_keys:
                    list_keys.append(key)
        list_keys.append('puntaje')
        list_keys.append('calidad_general')
        list_keys.append('mensaje')
        return list_keys

    def _get_user_name(self, user_id):
        user = get_collection('users', {'_id': ObjectId(user_id)})
        if user:
            return f'{user[0]["name"]} {user[0]["last_name"]}'
        return ''
